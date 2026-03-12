"""
Excel import/export — reads accounts from .xlsx and exports account data.

Uses openpyxl for Excel file handling.
"""

from __future__ import annotations

import json
import logging
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from bybit_manager.manager import Manager
from bybit_manager.scripts._excel_columns import (
    COLUMNS,
    EXPORT_COLUMNS,
    EMAIL_IMAP_ADDRESS_COLUMN,
    EMAIL_IMAP_PASSWORD_COLUMN,
)

logger = logging.getLogger("bybit_manager.scripts.excel")


async def import_accounts_from_excel(
    manager: Manager,
    file_path: str,
    group_name: str = "no_group",
    sheet_name: Optional[str] = None,
) -> Dict[str, Any]:
    """Import accounts from an Excel file.

    Expected columns match COLUMNS in _excel_columns.py.
    First row is header. email_address is required.

    Returns:
        {"imported": int, "skipped": int, "errors": list}
    """
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("openpyxl is required for Excel import: pip install openpyxl")

    wb = openpyxl.load_workbook(file_path, read_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    # Read header row
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"imported": 0, "skipped": 0, "errors": ["Empty file"]}

    header = [str(h).strip().lower() if h else "" for h in rows[0]]

    # Build column index map
    col_map: Dict[str, int] = {}
    for idx, name in enumerate(header):
        col_map[name] = idx

    if "email_address" not in col_map and "email" not in col_map:
        return {"imported": 0, "skipped": 0, "errors": ["Missing email_address column"]}

    email_idx = col_map.get("email_address", col_map.get("email", -1))

    imported = 0
    skipped = 0
    errors: List[str] = []

    for row_num, row in enumerate(rows[1:], start=2):
        try:
            if not row or email_idx >= len(row) or not row[email_idx]:
                skipped += 1
                continue

            email_address = str(row[email_idx]).strip()
            if not email_address or "@" not in email_address:
                errors.append(f"Row {row_num}: invalid email '{email_address}'")
                continue

            # Build account fields
            account_fields: Dict[str, Any] = {
                "email_address": email_address,
                "group_name": group_name,
            }
            imap_address = None
            imap_password = None

            for col_header, field_name, _ in COLUMNS:
                if col_header in col_map and field_name != "email_address":
                    idx = col_map[col_header]
                    if idx < len(row) and row[idx] is not None:
                        value = str(row[idx]).strip()
                        if value:
                            if field_name == EMAIL_IMAP_ADDRESS_COLUMN:
                                imap_address = value
                            elif field_name == EMAIL_IMAP_PASSWORD_COLUMN:
                                imap_password = value
                            else:
                                account_fields[field_name] = value

            # Handle cookies column (JSON)
            if "cookies" in col_map:
                idx = col_map["cookies"]
                if idx < len(row) and row[idx]:
                    try:
                        account_fields["cookies"] = json.loads(str(row[idx]))
                    except json.JSONDecodeError:
                        pass

            await manager.create_account(
                imap_address=imap_address,
                imap_password=imap_password,
                **account_fields,
            )
            imported += 1

        except Exception as e:
            errors.append(f"Row {row_num}: {e}")

    wb.close()
    logger.info("Excel import: %d imported, %d skipped, %d errors", imported, skipped, len(errors))
    return {"imported": imported, "skipped": skipped, "errors": errors}


async def export_accounts_to_excel(
    manager: Manager,
    file_path: str,
    group_name: Optional[str] = None,
) -> int:
    """Export accounts to an Excel file.

    Returns:
        Number of accounts exported.
    """
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("openpyxl is required for Excel export: pip install openpyxl")

    # Fetch all accounts
    accounts, total = await manager.get_accounts(
        group_name=group_name,
        page=1,
        page_size=100000,
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Accounts"

    # Write header
    headers = [col_header for col_header, _ in EXPORT_COLUMNS]
    ws.append(headers)

    # Write data rows
    for account in accounts:
        row = []
        for col_header, field_name in EXPORT_COLUMNS:
            if field_name in (EMAIL_IMAP_ADDRESS_COLUMN, EMAIL_IMAP_PASSWORD_COLUMN):
                # These come from the Email relationship
                email_obj = account.email
                value = getattr(email_obj, field_name, None) if email_obj else None
            elif field_name == "cookies":
                cookies = getattr(account, "cookies", None)
                value = json.dumps(cookies) if cookies else None
            elif field_name == "kyc_status":
                ks = getattr(account, "kyc_status", None)
                value = ks.value if hasattr(ks, "value") else ks
            else:
                value = getattr(account, field_name, None)

            row.append(value)
        ws.append(row)

    wb.save(file_path)
    wb.close()

    logger.info("Exported %d accounts to %s", len(accounts), file_path)
    return len(accounts)
